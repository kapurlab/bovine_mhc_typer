#!/usr/bin/env python3
"""Export per-run sample_sheet.csv files from the master multisheet Excel.

Each tab (one sequencing run) -> a sample_sheet.csv written into the matching
run folder in the library, matched by the 8-digit date in the tab name.
Columns: barcode, sample_id, amplicon, tissue. Barcodes are normalized to
barcodeNN; amplicons to canonical tokens (Bov7/11, BosEx, DRB3, 5'UTR).

  export_sheets_from_master.py <master.xlsx> <runs_library_dir>

Run it whenever the master Excel changes; the app reads the in-run sheets.
"""
import csv
import re
import sys
from pathlib import Path

import openpyxl


def norm_amplicon(a: str) -> str:
    low = re.sub(r"[()\s]", "", str(a or "")).lower()
    if "bov" in low:
        return "Bov7/11"
    if "bosex" in low:
        return "BosEx"
    if "drb3" in low:
        return "DRB3"
    if "utr" in low:
        return "5'UTR"
    return re.sub(r"[()]", "", str(a or "")).strip()


def main():
    master, libdir = sys.argv[1], Path(sys.argv[2])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    lib_by_date = {}
    for d in libdir.iterdir():
        m = re.match(r"(\d{8})", d.name)
        if d.is_dir() and m:
            lib_by_date[m.group(1)] = d

    for sn in wb.sheetnames:
        m = re.search(r"(\d{8})", sn)
        if not m:
            continue
        run = lib_by_date.get(m.group(1))
        if not run:
            print(f"  {sn}: no run folder for {m.group(1)}")
            continue
        tissue = "liver" if "liver" in sn.lower() else ("blood" if "blood" in sn.lower() else "")
        rows = [r for r in wb[sn].iter_rows(values_only=True) if r and r[0]]
        if len(rows) < 2:
            continue
        hdr = [str(c).strip().lower() if c else "" for c in rows[0]]

        def col(key):
            return next((i for i, h in enumerate(hdr) if key in h), None)

        ci_s, ci_a, ci_b = col("sample id"), col("amplicon"), col("barcode")
        if ci_s is None or ci_b is None:
            print(f"  {sn}: missing sample/barcode column")
            continue
        out = []
        for r in rows[1:]:
            bc = r[ci_b]
            if bc is None or str(bc).strip() == "":
                continue
            try:
                bcn = f"barcode{int(float(bc)):02d}"
            except (ValueError, TypeError):
                bcn = str(bc).strip()
            samp = str(r[ci_s]).strip() if r[ci_s] is not None else ""
            amp = norm_amplicon(r[ci_a]) if ci_a is not None else ""
            # animal_id defaults to sample_id (correct where sample_id is the
            # ear-tag, e.g. blood runs); edit it for liver runs where sample_id
            # is a per-sample label so genotypes group at the animal level.
            out.append([bcn, samp, samp, amp, tissue])
        (run / "sample_sheet.tsv").unlink(missing_ok=True)  # drop any stale sheet
        with open(run / "sample_sheet.csv", "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["barcode", "sample_id", "animal_id", "amplicon", "tissue"])
            w.writerows(out)
        print(f"  {sn} -> {run.name}/sample_sheet.csv ({len(out)} rows)")


if __name__ == "__main__":
    main()
